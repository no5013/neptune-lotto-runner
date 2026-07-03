#!/usr/bin/env python3
"""Export lottery winners to a print-ready Excel file."""

import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent


def parse_args():
    parser = argparse.ArgumentParser(description="Export lottery winners to Excel.")
    parser.add_argument(
        "--results",
        default=None,
        help="Path to results.json. Defaults to the latest run in output/.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output .xlsx path. Defaults to winners_export.xlsx in the results folder.",
    )
    return parser.parse_args()


def latest_results() -> Path:
    runs = sorted((BASE_DIR / "output").glob("run_*/results.json"))
    if not runs:
        raise FileNotFoundError("No run output found in output/")
    return runs[-1]



def thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def build_excel(winners: list, prices: dict, out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Winners"

    # ── Palette ──────────────────────────────────────────────────────────────
    TEAL       = "2A9D8F"
    TEAL_LIGHT = "E8F5F4"
    AMBER      = "F4A261"
    WHITE      = "FFFFFF"
    GRAY_ROW   = "F8F8F8"

    header_font  = Font(bold=True, color=WHITE, size=11)
    header_fill  = PatternFill("solid", fgColor=TEAL)
    total_font   = Font(bold=True, color="FFFFFF", size=11)
    total_fill   = PatternFill("solid", fgColor=AMBER)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left         = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = ["คิวที่", "ชื่อ", "LINE ID", "Email", "รอบรับของ", "ของที่ได้", "ราคา/ชิ้น", "ยอดรวม", "✓"]
    col_widths = [8, 18, 18, 30, 16, 32, 12, 12, 6]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = center
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────────────────────
    # Sort by queue number (None last)
    sorted_winners = sorted(winners, key=lambda p: p.get("queueNumber") or 9999)

    row = 2
    for i, person in enumerate(sorted_winners):
        items      = person.get("items", [])
        queue_num  = person.get("queueNumber")
        period     = person.get("pickupPeriod")
        period_str = f"{period['label']} ({period['time']} น.)" if period else ""
        total      = sum(prices.get(r["item"], 0) for r in items)
        item_names = "\n".join(f"• {r['item']}" for r in items)
        item_prices_str = "\n".join(
            f"฿{prices.get(r['item'], 0):,}" for r in items
        )

        fill_color = WHITE if i % 2 == 0 else GRAY_ROW
        row_fill = PatternFill("solid", fgColor=fill_color)

        values = [
            queue_num,
            person.get("name", ""),
            person.get("lineId", ""),
            person.get("email", ""),
            period_str,
            item_names,
            item_prices_str,
            total,
            "",  # checkbox column
        ]
        alignments = [center, left, left, left, center, left, center, center, center]

        for ci, (val, align) in enumerate(zip(values, alignments), start=1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.alignment = align
            cell.border    = thin_border()
            cell.fill      = row_fill

        # Highlight total cell
        total_cell = ws.cell(row=row, column=8)
        total_cell.font   = Font(bold=True, size=11)
        total_cell.fill   = PatternFill("solid", fgColor=TEAL_LIGHT)

        # Auto-height: roughly 15pt per item line
        ws.row_dimensions[row].height = max(18, len(items) * 18)
        row += 1

    # ── Summary row ───────────────────────────────────────────────────────────
    grand_total = sum(
        prices.get(r["item"], 0)
        for p in sorted_winners
        for r in p.get("items", [])
    )
    ws.cell(row=row, column=7, value="Grand Total").font = total_font
    ws.cell(row=row, column=7).fill                     = total_fill
    ws.cell(row=row, column=7).alignment                = center
    ws.cell(row=row, column=8, value=grand_total).font  = total_font
    ws.cell(row=row, column=8).fill                     = total_fill
    ws.cell(row=row, column=8).alignment                = center
    ws.row_dimensions[row].height = 22

    # ── Print settings ────────────────────────────────────────────────────────
    ws.page_setup.orientation   = "landscape"
    ws.page_setup.fitToWidth    = 1
    ws.page_setup.fitToHeight   = 0
    ws.print_title_rows         = "1:1"
    ws.sheet_view.showGridLines = True

    wb.save(out_path)


def main():
    args = parse_args()

    results_path = Path(args.results) if args.results else latest_results()
    with results_path.open(encoding="utf-8") as f:
        data = json.load(f)

    winners = [p for p in data.get("results", []) if p.get("items")]
    prices  = data.get("prices", {})

    out_path = Path(args.output) if args.output else results_path.parent / "winners_export.xlsx"
    build_excel(winners, prices, out_path)

    print(f"Exported {len(winners)} winners → {out_path}")
    print(f"Grand total : ฿{sum(prices.get(r['item'], 0) for p in winners for r in p['items']):,}")


if __name__ == "__main__":
    main()
